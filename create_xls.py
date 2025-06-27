#!/usr/bin/env python3
## begin license ##
#
# Gemeenschappelijke Metadata Harvester (GMH) uitbreidingen
#
# Copyright (C) 2025 Koninklijke Bibliotheek (KB) https://www.kb.nl
# Copyright (C) 2025 Seecr (Seek You Too B.V.) https://seecr.nl
#
# This file is part of "GMH-Harvester-Addon"
#
# "GMH-Harvester-Addon" is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# "GMH-Harvester-Addon" is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with "GMH-Harvester-Addon"; if not, write to the Free Software
# Foundation, Inc., 51 Franklin St, Fifth Floor, Boston, MA  02110-1301  USA
#
## end license ##

import argparse
from urllib.parse import urlencode

from pathlib import Path

from openpyxl import Workbook

from meresco.harvester.repositorystatus import RepositoryStatus
from meresco.harvester.harvesterdata import HarvesterData

import json


def write_nr_xls_to_gustos(domain_id, xls_count, gustos_count_path):
    if not gustos_count_path:
        return
    gustos_count_path.parent.mkdir(parents=True, exist_ok=True)
    current_data = {}
    if gustos_count_path.is_file():
        current_data = json.loads(gustos_count_path.read_text())
    xls = current_data.setdefault("GMH Harvester XLS", {})
    domain_info = xls.setdefault(f"Domain {domain_id}", {})
    prev_run_count = domain_info.get("nr_of_runs", {"count": 0})["count"]
    domain_info["nr_of_runs"] = {"count": prev_run_count + 1}
    domain_info["nr_of_xls"] = {"count": xls_count}
    tmp = gustos_count_path.with_suffix(".tmp")
    tmp.write_text(json.dumps(current_data))
    tmp.rename(gustos_count_path)


def main(state_path, log_path, data_path, target_path, domain_id, gustos_count_path):
    harvester_data = HarvesterData(dataPath=data_path)
    repository_status = RepositoryStatus(logPath=log_path, statePath=state_path)
    repository_status.addObserver(harvester_data)

    xls_count = 0
    for status in repository_status.getStatus(domainId=domain_id):
        xls_count += 1
        repository_id = status["repositoryId"]
        invalid_record_ids = list(
            repository_status.invalidRecords(domain_id, repository_id)
        )

        repository_data = harvester_data.getRepository(repository_id, domain_id)
        if len(invalid_record_ids) > 0:

            # Create workbook
            wb = Workbook()
            ws = wb.worksheets[0]

            # max. string length for worksheet title...
            ws.title = f"{repository_id} Invald OAI-PMH records"[:30]

            # Add content:
            max_col = 0
            for idx, recId in enumerate(invalid_record_ids):
                record_id = recId.split(":", 1)[-1]
                if len(record_id) > max_col:
                    max_col = len(record_id)
                etree = repository_status.getInvalidRecord(
                    domain_id, repository_id, record_id
                )

                diagnostic = etree.xpath(
                    "//diag:diagnostic/diag:details/text()",
                    namespaces={"diag": "http://www.loc.gov/zing/srw/diagnostic/"},
                )
                get_record_url = "{}?{}".format(
                    repository_data["baseurl"],
                    urlencode(
                        dict(
                            verb="GetRecord",
                            identifier=record_id,
                            metadataPrefix=repository_data["metadataPrefix"],
                        )
                    ),
                )
                firstcell = ws.cell(column=1, row=(idx + 1), value=record_id)
                firstcell.hyperlink = get_record_url
                ws.cell(column=2, row=(idx + 1), value=diagnostic[0])

            if max_col > 0:
                ws.column_dimensions["A"].width = max_col

            filename = target_path / f"{domain_id}_{repository_id}.xlsx"
            tmp_filename = filename.with_suffix(".tmp")
            try:
                wb.save(tmp_filename)
            finally:
                tmp_filename.rename(filename)

    write_nr_xls_to_gustos(domain_id, xls_count, gustos_count_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        prog="GMH XLS Report", description="Generate XLS files with status information"
    )
    parser.add_argument(
        "--state-path", type=Path, required=True, help="Path to harvester state"
    )
    parser.add_argument(
        "--log-path", type=Path, required=True, help="Path to harvester log"
    )
    parser.add_argument(
        "--data-path", type=Path, required=True, help="Path to harvester data"
    )
    parser.add_argument(
        "--target-path", type=Path, required=True, help="Path to store XLS files"
    )
    parser.add_argument(
        "--domain-id", type=str, required=True, help="Domain to generate XLS files for"
    )
    parser.add_argument(
        "--gustos-count-path", type=Path, help="File path to store Gustos count"
    )
    args = parser.parse_args()

    main(**vars(args))
